import pandas as pd
from openpyxl import load_workbook

def get_sheetnames_xlsx(filepath):
    wb = load_workbook(filepath, read_only=True, keep_links=False)
    return wb.sheetnames


def checkList(list):
    # funkcja sprawdzająca czy wszystkie elementy w podanej liście są takie same
    ele = list[0]
    chk = True
    for i in list:
        if ele != i:
            chk = False
            break;
    return chk


def table_maker(at, sheetname):
    # wczytanie pliku, modyfikacja i zapis do pliku
    df = pd.read_excel('Tab.xlsx', sheet_name=sheetname, header=None, engine='openpyxl')

    # usunięcie wierszy bez godzin odjazdu
    for ind, row in enumerate(df.values[1:]):
        lst = [str(j) for j in row[4:]]
        if checkList(lst):
            df.drop(ind+1, inplace=True)

    # nazwy przysanków z uchwały
    for num in range(len(df[0])):
        for nr in range(len(at[4])):
            if df.iloc[num, 0] == at.iloc[nr, 1]:
                df.iloc[num, 1] = at.iloc[nr, 0]

    # zmiana nazwy określonych komórek
    df.iloc[0, 0] = 'nr słupka'
    df.iloc[0, 1] = 'Przystanek'
    df.iloc[1, 2] = ''
    df.iloc[1, 3] = ''

    # ustawienie tytułów tabelek
    row_count = len(df[0]) - 1
    if sheetname == 'R1' or sheetname == 'R2':
        title = pd.DataFrame({0: [f'{df.iloc[1, 1]} - {df.iloc[row_count, 1]} - DNI ROBOCZE', ' ']}, index=[0, 1])
    elif sheetname == 'S1' or sheetname == 'S2':
        title = pd.DataFrame({0: [f'{df.iloc[1, 1]} - {df.iloc[row_count, 1]} - SOBOTY', ' ']}, index=[0, 1])
    elif sheetname == 'N1' or sheetname == 'N2':
        title = pd.DataFrame({0: [f'{df.iloc[1, 1]} - {df.iloc[row_count, 1]} - NIEDZIELE', ' ']}, index=[0, 1])
    else:
        title = pd.DataFrame({0: ['']}, index=[0])

    # połączenie tytułu z tabelką
    df = pd.concat([title, df]).reset_index(drop=True)
    return df


def row_len(data,row):
    # funkcja sumująca długość podanego wiersza
    list = []
    for i in range(len(data.columns)):
        if type(data.iloc[row,i]) != float:
            list.append(data.iloc[row,i])
    return len(list)

